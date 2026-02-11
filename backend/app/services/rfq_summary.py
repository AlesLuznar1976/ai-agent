"""
RFQ Summary Service - generacija sumarnega dokumenta za povpraševanja.

Zbere podatke iz emailov, prilog (PDF tekst, BOM, Gerber)
in generira sumarni PDF z AI povzetkom.
"""

import os
import json
from datetime import datetime
from typing import Optional

from sqlalchemy.orm import Session

from app.config import get_settings
from app.crud import emaili as crud_emaili, dokumenti as crud_dokumenti

settings = get_settings()

DOCUMENTS_BASE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
    "data", "documents"
)


async def generate_rfq_summary(
    db: Session,
    projekt_id: int,
    email_id: Optional[int] = None,
) -> dict:
    """Generiraj sumarni RFQ dokument za projekt.

    Args:
        db: SQLAlchemy session
        projekt_id: ID projekta
        email_id: Opcijsko - ID specifičnega emaila

    Returns:
        dict z pot do datoteke in statusom
    """
    from app.services.attachment_processor import extract_pdf_text

    # 1. Pridobi projekt
    from app.crud import projekti as crud_projekti
    projekt = crud_projekti.get_projekt_by_id(db, projekt_id)
    if not projekt:
        return {"error": f"Projekt {projekt_id} ne obstaja"}

    # 2. Pridobi povezane emaile
    emails = crud_emaili.list_emaili(db, projekt_id=projekt_id)
    if email_id:
        specific_email = crud_emaili.get_email_by_id(db, email_id)
        if specific_email and specific_email not in emails:
            emails.insert(0, specific_email)

    # 3. Zberi podatke iz emailov
    email_contents = []
    all_izvleceni = {}
    attachment_info = {"bom_files": [], "gerber_files": [], "pdf_files": [], "other_files": []}

    for email in emails:
        izvleceni = {}
        if email.izvleceni_podatki:
            try:
                izvleceni = json.loads(email.izvleceni_podatki)
            except (json.JSONDecodeError, TypeError):
                pass

        # Zberi izvlečene podatke (kasnejši emaili prepišejo zgodnejše)
        for k, v in izvleceni.items():
            if v and k not in ("kategorija", "zaupanje"):
                all_izvleceni[k] = v

        email_contents.append({
            "zadeva": email.zadeva,
            "posiljatelj": email.posiljatelj,
            "datum": email.datum.isoformat() if email.datum else "",
            "kategorija": email.kategorija,
            "telo": (email.telo or "")[:1000],
            "povzetek": izvleceni.get("povzetek", ""),
        })

        # Klasificiraj priloge
        if email.priloge:
            try:
                priloge = json.loads(email.priloge)
                for p in priloge:
                    if not isinstance(p, dict):
                        continue
                    name = p.get("name", "").lower()
                    tip = p.get("tip", "Drugo")
                    info = {"name": p.get("name", ""), "size": p.get("size", 0), "local_path": p.get("local_path", "")}

                    if tip == "BOM" or "bom" in name:
                        attachment_info["bom_files"].append(info)
                    elif tip == "Gerber" or "gerber" in name:
                        attachment_info["gerber_files"].append(info)
                    elif name.endswith(".pdf"):
                        attachment_info["pdf_files"].append(info)
                    else:
                        attachment_info["other_files"].append(info)
            except (json.JSONDecodeError, TypeError):
                pass

    # 4. Izvleči tekst iz PDF prilog
    pdf_texts = []
    for pdf_info in attachment_info["pdf_files"]:
        local_path = pdf_info.get("local_path", "")
        if local_path and os.path.exists(local_path):
            try:
                with open(local_path, "rb") as f:
                    pdf_bytes = f.read()
                text = extract_pdf_text(pdf_bytes)
                if text:
                    pdf_texts.append({"name": pdf_info["name"], "text": text[:3000]})
            except Exception:
                pass

    # 5. Preberi BOM datoteke
    bom_summaries = []
    for bom_info in attachment_info["bom_files"]:
        local_path = bom_info.get("local_path", "")
        if local_path and os.path.exists(local_path):
            try:
                bom_summary = _read_bom_summary(local_path)
                bom_summaries.append({"name": bom_info["name"], **bom_summary})
            except Exception:
                pass

    # 6. Generiraj AI povzetek
    ai_summary = await _generate_ai_summary(
        email_contents=email_contents,
        izvleceni=all_izvleceni,
        pdf_texts=pdf_texts,
        bom_summaries=bom_summaries,
        attachment_info=attachment_info,
    )

    # 7. Generiraj PDF
    stevilka = projekt.stevilka_projekta if hasattr(projekt, "stevilka_projekta") else f"PRJ-{projekt_id}"

    # Pridobi ime stranke
    stranka_ime = all_izvleceni.get("stranka", "")
    if not stranka_ime and emails:
        sender = emails[0].posiljatelj or ""
        if "<" in sender:
            stranka_ime = sender.split("<")[0].strip()
        else:
            stranka_ime = sender

    kontakt_email = ""
    if emails:
        sender = emails[0].posiljatelj or ""
        if "<" in sender and ">" in sender:
            kontakt_email = sender.split("<")[1].split(">")[0]
        else:
            kontakt_email = sender

    pdf_path = _generate_pdf(
        projekt_id=projekt_id,
        stevilka=stevilka,
        stranka=stranka_ime,
        kontakt=kontakt_email,
        ai_summary=ai_summary,
        attachment_info=attachment_info,
        bom_summaries=bom_summaries,
        izvleceni=all_izvleceni,
    )

    # 8. Registriraj v DB
    filename = os.path.basename(pdf_path)
    db_dok = crud_dokumenti.create_dokument(
        db,
        projekt_id=projekt_id,
        naziv_datoteke=filename,
        pot_do_datoteke=pdf_path,
        tip="TIV",
    )

    return {
        "message": f"RFQ Summary generiran: {filename}",
        "path": pdf_path,
        "dokument_id": db_dok.id,
        "stevilka_projekta": stevilka,
    }


def _read_bom_summary(file_path: str) -> dict:
    """Preberi BOM iz Excel in vrne povzetek."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(max_row=200, values_only=True))
        wb.close()

        if not rows:
            return {"components": 0, "preview": []}

        # Prvih 5 vrstic za preview
        header = [str(c) if c else "" for c in rows[0]] if rows else []
        preview = []
        for row in rows[1:6]:
            preview.append([str(c) if c else "" for c in row])

        return {
            "components": len(rows) - 1,  # minus header
            "header": header,
            "preview": preview,
        }
    except Exception as e:
        return {"components": 0, "error": str(e)}


async def _generate_ai_summary(
    email_contents: list[dict],
    izvleceni: dict,
    pdf_texts: list[dict],
    bom_summaries: list[dict],
    attachment_info: dict,
) -> str:
    """Generiraj AI povzetek iz vseh virov."""
    from app.llm import get_llm_router, TaskType

    llm = get_llm_router()

    # Pripravi kontekst
    context_parts = []

    # Email vsebine
    for ec in email_contents[:5]:
        context_parts.append(f"Email: {ec['zadeva']}\nOd: {ec['posiljatelj']}\n{ec['telo'][:500]}")

    # Izvlečeni podatki
    if izvleceni:
        context_parts.append(f"Izvlečeni podatki: {json.dumps(izvleceni, ensure_ascii=False)}")

    # PDF teksti
    for pt in pdf_texts[:3]:
        context_parts.append(f"PDF '{pt['name']}':\n{pt['text'][:1500]}")

    # BOM povzetki
    for bs in bom_summaries:
        context_parts.append(f"BOM '{bs.get('name', '?')}': {bs.get('components', 0)} komponent")

    context = "\n\n---\n\n".join(context_parts)

    prompt = f"""Na podlagi naslednjih informacij o povpraševanju za PCB/SMT izdelavo napiši kratek poslovni povzetek.

INFORMACIJE:
{context}

NAVODILA:
- Piši v slovenščini
- Bodi jedrnat (3-5 stavkov)
- Omeni: stranko, kaj želijo, količine, posebne zahteve
- Če informacija ni na voljo, jo izpusti

POVZETEK:"""

    try:
        summary = await llm.complete(
            prompt,
            task_type=TaskType.DOCUMENT_GENERATION,
            contains_sensitive=False,
        )
        return summary.strip()
    except Exception as e:
        print(f"AI summary error: {e}")
        # Fallback - sestavi iz izvlečenih podatkov
        parts = []
        if izvleceni.get("stranka"):
            parts.append(f"Povpraševanje od {izvleceni['stranka']}.")
        if izvleceni.get("povzetek"):
            parts.append(izvleceni["povzetek"])
        return " ".join(parts) if parts else "Povzetek ni na voljo."


def _generate_pdf(
    projekt_id: int,
    stevilka: str,
    stranka: str,
    kontakt: str,
    ai_summary: str,
    attachment_info: dict,
    bom_summaries: list[dict],
    izvleceni: dict,
) -> str:
    """Generiraj RFQ Summary PDF z fpdf2."""
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Header
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "LUZNAR ELECTRONICS d.o.o.", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 5, "", new_x="LMARGIN", new_y="NEXT")  # spacing

    # Horizontalna črta
    pdf.set_draw_color(0, 0, 0)
    pdf.set_line_width(0.5)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.cell(0, 5, "", new_x="LMARGIN", new_y="NEXT")

    # Naslov
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "POVZETEK POVPRASEVANJA (RFQ Summary)", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.cell(0, 5, "", new_x="LMARGIN", new_y="NEXT")

    # Metadata
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(40, 7, "Projekt:", new_x="RIGHT")
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 7, stevilka, new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(40, 7, "Datum:", new_x="RIGHT")
    pdf.cell(0, 7, datetime.now().strftime("%Y-%m-%d"), new_x="LMARGIN", new_y="NEXT")

    pdf.cell(40, 7, "Stranka:", new_x="RIGHT")
    pdf.cell(0, 7, _safe_latin(stranka), new_x="LMARGIN", new_y="NEXT")

    pdf.cell(40, 7, "Kontakt:", new_x="RIGHT")
    pdf.cell(0, 7, _safe_latin(kontakt), new_x="LMARGIN", new_y="NEXT")

    pdf.cell(0, 5, "", new_x="LMARGIN", new_y="NEXT")

    # 1. POVZETEK ZAHTEVE
    _pdf_section(pdf, "1. POVZETEK ZAHTEVE")
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 6, _safe_latin(ai_summary))
    pdf.cell(0, 5, "", new_x="LMARGIN", new_y="NEXT")

    # 2. PREJETI DOKUMENTI
    _pdf_section(pdf, "2. PREJETI DOKUMENTI")
    pdf.set_font("Helvetica", "", 10)

    for bom in attachment_info.get("bom_files", []):
        count = ""
        for bs in bom_summaries:
            if bs.get("name") == bom.get("name"):
                count = f" ({bs.get('components', '?')} komponent)"
        pdf.cell(0, 6, f"  - BOM: {_safe_latin(bom.get('name', '?'))}{count}", new_x="LMARGIN", new_y="NEXT")

    for gerber in attachment_info.get("gerber_files", []):
        size_kb = gerber.get("size", 0) / 1024
        pdf.cell(0, 6, f"  - Gerber: {_safe_latin(gerber.get('name', '?'))} ({size_kb:.0f} KB)", new_x="LMARGIN", new_y="NEXT")

    for pdf_file in attachment_info.get("pdf_files", []):
        pdf.cell(0, 6, f"  - PDF: {_safe_latin(pdf_file.get('name', '?'))}", new_x="LMARGIN", new_y="NEXT")

    for other in attachment_info.get("other_files", []):
        pdf.cell(0, 6, f"  - Drugo: {_safe_latin(other.get('name', '?'))}", new_x="LMARGIN", new_y="NEXT")

    if not any(attachment_info.values()):
        pdf.cell(0, 6, "  Ni prilog.", new_x="LMARGIN", new_y="NEXT")

    pdf.cell(0, 5, "", new_x="LMARGIN", new_y="NEXT")

    # 3. KLJUCNI PODATKI
    _pdf_section(pdf, "3. KLJUCNI PODATKI")
    pdf.set_font("Helvetica", "", 10)

    kolicina = izvleceni.get("kolicina") or izvleceni.get("quantity") or "ni podatka"
    tip_vezja = izvleceni.get("tip_projekta") or izvleceni.get("tip_vezja") or "ni podatka"
    posebne = izvleceni.get("posebne_zahteve") or izvleceni.get("special_requirements") or "ni podatka"

    pdf.cell(0, 6, f"  - Kolicina: {_safe_latin(str(kolicina))}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"  - Tip vezja: {_safe_latin(str(tip_vezja))}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, f"  - Posebne zahteve: {_safe_latin(str(posebne))}", new_x="LMARGIN", new_y="NEXT")

    pdf.cell(0, 5, "", new_x="LMARGIN", new_y="NEXT")

    # 4. NASLEDNJI KORAKI
    _pdf_section(pdf, "4. NASLEDNJI KORAKI")
    pdf.set_font("Helvetica", "", 10)
    steps = [
        "[ ] Pregled BOM-a",
        "[ ] Vnos v CalcuQuote",
        "[ ] Priprava ponudbe",
    ]
    for step in steps:
        pdf.cell(0, 6, f"  {step}", new_x="LMARGIN", new_y="NEXT")

    # Shrani
    project_dir = os.path.join(DOCUMENTS_BASE_PATH, str(projekt_id))
    os.makedirs(project_dir, exist_ok=True)

    filename = f"RFQ_Summary_{stevilka}.pdf"
    filepath = os.path.join(project_dir, filename)
    pdf.output(filepath)

    return filepath


def _pdf_section(pdf, title: str):
    """Dodaj naslov sekcije v PDF."""
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, title, new_x="LMARGIN", new_y="NEXT")


def _safe_latin(text: str) -> str:
    """Pretvori tekst v latin-1 safe obliko za fpdf2 (brez UTF-8 fontov)."""
    if not text:
        return ""
    # Zamenjaj znake ki jih latin-1 ne podpira
    replacements = {
        "\u0161": "s", "\u0160": "S",  # š, Š
        "\u010d": "c", "\u010c": "C",  # č, Č
        "\u017e": "z", "\u017d": "Z",  # ž, Ž
        "\u0107": "c", "\u0106": "C",  # ć, Ć
        "\u0111": "d", "\u0110": "D",  # đ, Đ
        "\u2013": "-", "\u2014": "-",  # em/en dash
        "\u2018": "'", "\u2019": "'",  # smart quotes
        "\u201c": '"', "\u201d": '"',
        "\u2026": "...",  # ellipsis
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)

    # Odstrani preostale ne-latin-1 znake
    return text.encode("latin-1", errors="replace").decode("latin-1")
